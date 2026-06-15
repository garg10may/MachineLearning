from __future__ import annotations

from dataclasses import dataclass
from datetime import date, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
from faker import Faker
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = PROJECT_ROOT / "data"
SEED = 42


SPECIALTIES = [
    "Cardiology",
    "Endocrinology",
    "Primary Care",
    "Pulmonology",
    "Oncology",
    "Neurology",
    "Rheumatology",
]

REGIONS = ["Northeast", "Midwest", "South", "West"]
PAYER_TYPES = ["Commercial", "Medicare", "Medicaid", "Cash"]
CHANNELS = ["In-person", "Phone", "Email", "Video", "Conference"]
CALL_TOPICS = [
    "Clinical evidence",
    "Patient access",
    "Prior authorization",
    "Dosing discussion",
    "New indication",
    "Adherence support",
]

DIAGNOSIS_CODES = {
    "Cardiology": ["I10", "I25.10", "I50.9", "E78.5"],
    "Endocrinology": ["E11.9", "E10.65", "E03.9", "E66.9"],
    "Primary Care": ["I10", "E11.9", "J45.909", "M54.50"],
    "Pulmonology": ["J45.909", "J44.9", "R06.02", "G47.33"],
    "Oncology": ["C50.919", "C34.90", "D64.9", "Z51.11"],
    "Neurology": ["G43.909", "G40.909", "G20", "R51.9"],
    "Rheumatology": ["M06.9", "M32.9", "M10.9", "M79.7"],
}

PROCEDURE_CODES = ["99213", "99214", "93000", "83036", "96413", "94010", "20610"]
MEDICATION_CLASSES = [
    "Antihypertensive",
    "Antidiabetic",
    "Bronchodilator",
    "Biologic",
    "Analgesic",
    "Statin",
    "Chemotherapy",
]


@dataclass(frozen=True)
class GenerationConfig:
    providers: int = 125
    claims: int = 5_000
    ehr_encounters: int = 3_000
    call_activities: int = 1_200
    patients: int = 2_000


def random_dates(
    rng: np.random.Generator,
    count: int,
    start: date,
    end: date,
) -> list[date]:
    days = (end - start).days
    offsets = rng.integers(0, days + 1, size=count)
    return [start + timedelta(days=int(offset)) for offset in offsets]


def weighted_choice(
    rng: np.random.Generator,
    values: list[str],
    count: int,
    weights: list[float] | None = None,
) -> np.ndarray:
    probabilities = None
    if weights is not None:
        probabilities = np.array(weights, dtype=float)
        probabilities = probabilities / probabilities.sum()
    return rng.choice(values, size=count, p=probabilities)


def build_provider_data(
    fake: Faker,
    rng: np.random.Generator,
    config: GenerationConfig,
) -> pd.DataFrame:
    provider_ids = [f"HCP{idx:04d}" for idx in range(1, config.providers + 1)]
    specialties = weighted_choice(
        rng,
        SPECIALTIES,
        config.providers,
        weights=[0.16, 0.14, 0.24, 0.1, 0.11, 0.11, 0.14],
    )
    regions = weighted_choice(rng, REGIONS, config.providers)
    years_in_practice = rng.integers(2, 36, size=config.providers)
    patient_panel_size = rng.integers(350, 3_200, size=config.providers)
    academic_affiliation = rng.choice([True, False], size=config.providers, p=[0.28, 0.72])
    digital_adoption_score = np.clip(rng.normal(62, 18, size=config.providers), 5, 100)
    access_restriction_score = np.clip(rng.normal(42, 21, size=config.providers), 0, 100)
    market_potential_score = np.clip(
        patient_panel_size / patient_panel_size.max() * 100 + rng.normal(0, 8, size=config.providers),
        0,
        100,
    )

    provider_data = pd.DataFrame(
        {
            "hcp_id": provider_ids,
            "provider_name": [fake.name() for _ in provider_ids],
            "specialty": specialties,
            "region": regions,
            "state": [fake.state_abbr() for _ in provider_ids],
            "practice_type": weighted_choice(
                rng,
                ["Independent", "Hospital-owned", "Academic", "Integrated Delivery Network"],
                config.providers,
                weights=[0.34, 0.31, 0.14, 0.21],
            ),
            "years_in_practice": years_in_practice,
            "patient_panel_size": patient_panel_size,
            "academic_affiliation": academic_affiliation,
            "digital_adoption_score": digital_adoption_score.round(1),
            "access_restriction_score": access_restriction_score.round(1),
            "market_potential_score": market_potential_score.round(1),
        }
    )

    provider_data["synthetic_segment"] = pd.cut(
        (
            0.45 * provider_data["market_potential_score"]
            + 0.35 * provider_data["digital_adoption_score"]
            - 0.2 * provider_data["access_restriction_score"]
        ),
        bins=[-100, 35, 55, 72, 120],
        labels=["Low Opportunity", "Nurture", "Growth", "Priority"],
    ).astype(str)
    return provider_data


def build_claims_data(
    rng: np.random.Generator,
    provider_data: pd.DataFrame,
    config: GenerationConfig,
) -> pd.DataFrame:
    hcp_ids = provider_data["hcp_id"].to_numpy()
    hcp_lookup = provider_data.set_index("hcp_id")["specialty"].to_dict()
    claim_hcps = rng.choice(hcp_ids, size=config.claims)
    claim_specialties = [hcp_lookup[hcp_id] for hcp_id in claim_hcps]
    service_dates = random_dates(rng, config.claims, date(2024, 1, 1), date(2025, 12, 31))
    claim_amount = rng.gamma(shape=2.2, scale=145, size=config.claims) + 75
    denied = rng.choice([True, False], size=config.claims, p=[0.09, 0.91])
    paid_ratio = np.where(denied, rng.uniform(0.0, 0.18, size=config.claims), rng.uniform(0.62, 0.96, size=config.claims))

    return pd.DataFrame(
        {
            "claim_id": [f"CLM{idx:07d}" for idx in range(1, config.claims + 1)],
            "patient_id": [f"PAT{idx:06d}" for idx in rng.integers(1, config.patients + 1, size=config.claims)],
            "hcp_id": claim_hcps,
            "service_date": service_dates,
            "diagnosis_code": [
                rng.choice(DIAGNOSIS_CODES[specialty]) for specialty in claim_specialties
            ],
            "procedure_code": rng.choice(PROCEDURE_CODES, size=config.claims),
            "payer_type": weighted_choice(
                rng,
                PAYER_TYPES,
                config.claims,
                weights=[0.48, 0.32, 0.16, 0.04],
            ),
            "claim_amount": claim_amount.round(2),
            "paid_amount": (claim_amount * paid_ratio).round(2),
            "days_to_pay": rng.integers(5, 75, size=config.claims),
            "claim_denied": denied,
            "prior_authorization_required": rng.choice(
                [True, False],
                size=config.claims,
                p=[0.22, 0.78],
            ),
        }
    )


def build_ehr_data(
    rng: np.random.Generator,
    provider_data: pd.DataFrame,
    config: GenerationConfig,
) -> pd.DataFrame:
    hcp_ids = provider_data["hcp_id"].to_numpy()
    encounter_dates = random_dates(rng, config.ehr_encounters, date(2024, 1, 1), date(2025, 12, 31))
    age = rng.integers(18, 91, size=config.ehr_encounters)
    bmi = np.clip(rng.normal(29, 6.5, size=config.ehr_encounters), 17, 55)
    blood_pressure_systolic = np.clip(rng.normal(132, 18, size=config.ehr_encounters), 90, 210)
    a1c = np.clip(rng.normal(7.2, 1.6, size=config.ehr_encounters), 4.8, 14.5)
    adherence_score = np.clip(rng.normal(71, 19, size=config.ehr_encounters), 0, 100)

    return pd.DataFrame(
        {
            "encounter_id": [f"ENC{idx:07d}" for idx in range(1, config.ehr_encounters + 1)],
            "patient_id": [f"PAT{idx:06d}" for idx in rng.integers(1, config.patients + 1, size=config.ehr_encounters)],
            "hcp_id": rng.choice(hcp_ids, size=config.ehr_encounters),
            "encounter_date": encounter_dates,
            "age": age,
            "sex": weighted_choice(rng, ["F", "M"], config.ehr_encounters, weights=[0.53, 0.47]),
            "bmi": bmi.round(1),
            "blood_pressure_systolic": blood_pressure_systolic.round(0).astype(int),
            "a1c_percent": a1c.round(1),
            "medication_class": rng.choice(MEDICATION_CLASSES, size=config.ehr_encounters),
            "adherence_score": adherence_score.round(1),
            "follow_up_recommended": (
                (a1c > 8.5)
                | (blood_pressure_systolic > 150)
                | (adherence_score < 55)
            ),
            "risk_tier": pd.cut(
                0.45 * (a1c - 5) + 0.02 * (blood_pressure_systolic - 110) + 0.03 * (bmi - 22),
                bins=[-10, 2, 4, 20],
                labels=["Low", "Medium", "High"],
            ).astype(str),
        }
    )


def build_call_activity(
    rng: np.random.Generator,
    provider_data: pd.DataFrame,
    config: GenerationConfig,
) -> pd.DataFrame:
    hcp_ids = provider_data["hcp_id"].to_numpy()
    call_dates = random_dates(rng, config.call_activities, date(2024, 1, 1), date(2025, 12, 31))
    channel = weighted_choice(
        rng,
        CHANNELS,
        config.call_activities,
        weights=[0.32, 0.22, 0.27, 0.13, 0.06],
    )

    return pd.DataFrame(
        {
            "activity_id": [f"ACT{idx:07d}" for idx in range(1, config.call_activities + 1)],
            "hcp_id": rng.choice(hcp_ids, size=config.call_activities),
            "activity_date": call_dates,
            "sales_rep_id": [f"REP{idx:03d}" for idx in rng.integers(1, 26, size=config.call_activities)],
            "channel": channel,
            "topic": rng.choice(CALL_TOPICS, size=config.call_activities),
            "duration_minutes": np.where(
                channel == "Email",
                rng.integers(1, 7, size=config.call_activities),
                rng.integers(8, 55, size=config.call_activities),
            ),
            "engagement_score": np.clip(rng.normal(66, 20, size=config.call_activities), 0, 100).round(1),
            "sentiment": weighted_choice(
                rng,
                ["Positive", "Neutral", "Negative"],
                config.call_activities,
                weights=[0.48, 0.42, 0.10],
            ),
            "sample_dropped": rng.choice([True, False], size=config.call_activities, p=[0.18, 0.82]),
            "follow_up_requested": rng.choice([True, False], size=config.call_activities, p=[0.31, 0.69]),
        }
    )


def build_codebook() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("claims_data", "claim_id", "Synthetic claim identifier"),
            ("claims_data", "patient_id", "Synthetic patient identifier"),
            ("claims_data", "hcp_id", "Provider identifier used to join provider_data"),
            ("claims_data", "diagnosis_code", "Synthetic ICD-10 diagnosis code"),
            ("ehr_data", "encounter_id", "Synthetic EHR encounter identifier"),
            ("ehr_data", "a1c_percent", "Synthetic HbA1c percentage"),
            ("ehr_data", "risk_tier", "Derived synthetic patient-level risk tier"),
            ("provider_data", "hcp_id", "Unique healthcare provider identifier"),
            ("provider_data", "synthetic_segment", "Rule-based label for demo segmentation"),
            ("call_activity", "activity_id", "Synthetic field activity identifier"),
            ("call_activity", "engagement_score", "Synthetic engagement score from 0 to 100"),
        ],
        columns=["dataset", "field", "description"],
    )


def write_outputs(
    claims_data: pd.DataFrame,
    ehr_data: pd.DataFrame,
    provider_data: pd.DataFrame,
    call_activity: pd.DataFrame,
) -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    datasets = {
        "claims_data": claims_data,
        "ehr_data": ehr_data,
        "provider_data": provider_data,
        "call_activity": call_activity,
    }

    for file_stem, data_frame in datasets.items():
        data_frame.to_csv(DATA_DIR / f"{file_stem}.csv", index=False)

    workbook_path = DATA_DIR / "hcp_segmentation_synthetic_data.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        for sheet_name, data_frame in datasets.items():
            data_frame.to_excel(writer, sheet_name=sheet_name, index=False)
            format_worksheet(writer.sheets[sheet_name], data_frame)
        build_codebook().to_excel(writer, sheet_name="codebook", index=False)
        format_worksheet(writer.sheets["codebook"], build_codebook())


def format_worksheet(worksheet, data_frame: pd.DataFrame) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for index, column_name in enumerate(data_frame.columns, start=1):
        column_values = data_frame[column_name].astype(str).head(200)
        max_length = max([len(str(column_name)), *column_values.map(len).tolist()])
        worksheet.column_dimensions[get_column_letter(index)].width = min(max(max_length + 2, 12), 32)


def main() -> None:
    config = GenerationConfig()
    fake = Faker("en_US")
    Faker.seed(SEED)
    rng = np.random.default_rng(SEED)

    provider_data = build_provider_data(fake, rng, config)
    claims_data = build_claims_data(rng, provider_data, config)
    ehr_data = build_ehr_data(rng, provider_data, config)
    call_activity = build_call_activity(rng, provider_data, config)

    write_outputs(
        claims_data=claims_data,
        ehr_data=ehr_data,
        provider_data=provider_data,
        call_activity=call_activity,
    )

    print(f"Wrote synthetic data to {DATA_DIR}")
    print(f"Providers: {len(provider_data):,}")
    print(f"Claims: {len(claims_data):,}")
    print(f"EHR encounters: {len(ehr_data):,}")
    print(f"Call activities: {len(call_activity):,}")


if __name__ == "__main__":
    main()
